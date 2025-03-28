import anvil.server
import pandas as pd
import doctest
from phonetics import metaphone, nysiis
from fuzzywuzzy import fuzz, process, utils
import math
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl.styles import Color, PatternFill, Font, Border
from copy import copy
from io import BytesIO

# ----- Pre-existing definitions from your Colab code -----

# Global variables for your mappings etc.
# (Ensure that MASTER_MAPPING, GEO_TERMS, etc. are defined elsewhere in your code,
#  or are defined in the portion of the file that remains unchanged.)
GEO_TERMS = {
    'north': 'Semeni',
    'south': 'Debubi',
    'east': 'Misirak\'i',
    'west': 'Mi\'irabi',
    'central': 'Ma\'ikelawi',
    'city': 'Ketema',
    'rural': 'Get\'eri',
    'valley': 'Shelek\'o',
    'river': 'Wenizi',
    'mountains': 'Terarochi',
    'desert': 'Bereha',
    'sea': 'Bahiri',
    'new': 'Adisi',
}

# reg_zone_woreda_file = "2024_ETH_RBLF_Admin_Names_Compiled copy.xlsx"

# # read by default 1st sheet of an excel file
# df = pd.read_excel(reg_zone_woreda_file, engine='openpyxl')

# MASTER_MAPPING = df['reg_zone_woreda'].values.tolist()

# REGION NAMES assuming structure in master_mapping is region--zone--woreda
"""
set containing the region names
NOTE: REGIONS is dynamically computed - change only the MASTER_MAPPING above
"""


# ZONE NAMES and WOREDA NAMES
"""
dictionary containing the woreda in each zone
NOTE: REGIONS_ZONES is dynamically computed - change only the MASTER_MAPPING above
NOTE: ZONES_WOREDAS is dynamically computed - change only the MASTER_MAPPING above
"""
# REGIONS_ZONES = defaultdict(list)
# for line in MASTER_MAPPING:
#     region, zone, woreda = line.split("_")
#     REGIONS_ZONES[region].append(zone)

# ZONES_WOREDAS = defaultdict(list)
# for line in MASTER_MAPPING:
#     region, zone, woreda = line.split("_")
#     ZONES_WOREDAS[zone].append(woreda)

# Processing Functions

def tr_geo_terms(s: str):

  return " ".join(
      GEO_TERMS.get(fragment.lower(), fragment) for fragment in s.split()
  )

def confidence_score(scores):

  # Set maxSim to the best similarity found
  maxSim = max(scores)

  # Calculate the denominator
  den = sum([math.exp(1 - math.pow(maxSim / score, 3)) for score in scores])

  # Calculate the confidence score for each score
  confidences = [math.exp(1 - math.pow(maxSim / score, 3)) / den for score in scores]
  return confidences

def custom_scorer(s1_raw: str, s2_raw: str) -> int:
  """
  return a measure of the sequences' similarity between 0 and 100, using different algorithms.

  passed as an argument to process.extract* functions in order to define a custom
  set of criteria for determining string similarity

  steps:
  1. process strings (remove whitespace, non-alpha, etc.) and check for corner cases
  2. replace geo terms if present
  3. compute baseline score using only fuzzy match
  4. compute similarity score using fuzzy on phonetic codes
  5. return a weighted average of phonetics and fuzzy

  areas for improvement: missing words in name,
  differentiating same name in different states
  """

  s1, s2 = tuple(map(utils.full_process, (s1_raw, s2_raw)))

  if not s1 or not s2:
    print(s1_raw, s2_raw)
    return 1 # returning 1 instead of 0 to not break confidence score algorithm

  s1, s2 = tr_geo_terms(s1), tr_geo_terms(s2) # 3. substitution of terms added 5 percent accuracy

  fuzzy_similarity = fuzz.token_sort_ratio(s1,s2) # 1. added 11 percent accuracy

  phonetic_similarity = fuzz.ratio(
      metaphone(s1), metaphone(s2)
  )

  # 2. weighted factor added 5%
  return int(fuzzy_similarity * 0.67 + phonetic_similarity * 0.33)

def get_prediction(inputs, standards):
  """
  given a list of inputs, returns a tuple-list of predictions and confidence score
  """

  output = []

  for locality in inputs:
    guesses_raw = process.extractBests(locality, standards, scorer=custom_scorer, limit=7)

    guesses = [ guess for guess in guesses_raw if guess[-1]]

    confidences = confidence_score(
        list(map(lambda x: x[-1], guesses))
    )

    guess = guesses[0][0]
    confidence = confidences[0]


    output.append(
        (guess, confidence)
    )

  return output

# ----- Anvil Server Module Function -----
# (No connection call is needed because this code is now running on Anvil servers)

@anvil.server.callable
def standardize(master_mapping_file, input_file, sheetname, output_file_name, low_confidence_score, xmed_confidence_score):

    # set default values if not specified in anvil
    sheetname = "NEMO" if not sheetname else sheetname
    master_mapping_file = "2024_ETH_RBLF_Admin_Names_Compiled copy.xlsx" if not master_mapping_file else master_mapping_file
    output_file_name = "Ethiopia_Mappings.xlsx" if not output_file_name else output_file_name
    low_confidence_score = 0.3 if not low_confidence_score else float(low_confidence_score)
    xmed_confidence_score = 0.4 if not xmed_confidence_score else float(xmed_confidence_score)

    # Convert the master mapping file (uploaded as an Anvil File object) to bytes
    mm_file_bytes = master_mapping_file.get_bytes()

    # Load the bytes into a BytesIO stream
    mm_file_stream = BytesIO(mm_file_bytes)

    # Load the Excel file into a workbook
    try:
      mm_workbook = load_workbook(mm_file_stream, data_only=True)
      # if "df" not in mm_dataframe.sheetnames
      mm_sheet = mm_workbook["df"]

      mm_data = mm_sheet.values
      mm_cols = next(mm_data)
      mm_df = pd.DataFrame(mm_data, columns=mm_cols)
      MASTER_MAPPING = mm_df['reg_zone_woreda'].values.tolist()

      # MASTER_MAPPING = mm_dataframe["df"]['reg_zone_woreda'].values.tolist()
    except Exception as e:
      return f"Error has occured: {e}"

    REGIONS = {region.split("_")[0] for region in MASTER_MAPPING}  # dynamic mapping

    REGIONS_ZONES = defaultdict(list)
    for line in MASTER_MAPPING:
        region, zone, woreda = line.split("_")
        REGIONS_ZONES[region].append(zone)
    
    ZONES_WOREDAS = defaultdict(list)
    for line in MASTER_MAPPING:
        region, zone, woreda = line.split("_")
        ZONES_WOREDAS[zone].append(woreda)

    # Convert the input file (uploaded as an Anvil File object) to bytes
    # ADD BACK IN
    input_file_bytes = input_file.get_bytes()

    # Load the bytes into a BytesIO stream
    input_file_stream = BytesIO(input_file_bytes)
    try:
      # Load the Excel file into a workbook
      dataframe = load_workbook(input_file_stream, data_only=True)
      # Find the sheet specified by the user
      # if sheetname not in dataframe.sheetnames:
      #     raise ValueError(f"Sheet {sheetname} not found.")

      # return sheetname
      sheet = dataframe[sheetname]
    except Exception as e:
      return f"Error has occured: {e}"

    # Initialize variables for processing
    header = True
    regions = []
    zones = []
    woredas = []

    # Iterate through the rows and extract region, zone, and woreda information
    for value in sheet.iter_rows(min_col=1, max_col=3, values_only=True):
        # Skip header row
        if header:
            header = False
            continue
        # Stop iterating when empty cells are reached
        if value[0] is None or value[1] is None or value[2] is None:
            break

        region = value[0]
        zone = value[1]
        woreda = value[2]  # Adjust if you want to clean or split woreda names

        # Add to lists
        regions.append(region)
        zones.append(zone)
        woredas.append(woreda)

    # format the header for the new column
    sheet.insert_cols(1, 1)
    column_number = 1
    column = str(chr(64 + column_number))
    sheet.column_dimensions[column].width = 30
    sheet.cell(row=1,column=1).font = copy(sheet['B1'].font)
    sheet.cell(row=1,column=1).fill = copy(sheet['B1'].fill)
    sheet['A1'] = 'Mappings'

    # color for cells with low confidence
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
    orange = PatternFill(start_color='FF8000', end_color='FF0000', fill_type = 'solid')

    index = 2
    print(f'length of woredas = {len(woredas)}')
    print(woredas)
    print('-----')
    for i in range(len(woredas)):
        region = regions[i]
        zone = zones[i]
        woreda = woredas[i]
        # Step 1: Get prediction for the region
        region_prediction = get_prediction([region], REGIONS)
        # Step 2: Get prediction for the zone based on the region's prediction
        zone_prediction = get_prediction([zone], REGIONS_ZONES[region_prediction[0][0]])
        # Step 3: Get prediction for the woreda based on the zone's prediction
        woreda_prediction = get_prediction([woreda], ZONES_WOREDAS[zone_prediction[0][0]])
        # Write the data to the cell
        sheet[f'A{index}'] = f'{region_prediction[0][0]}--{zone_prediction[0][0]}--{woreda_prediction[0][0]}'

        # Color the cell based on the confidence
        if woreda_prediction[0][1] < low_confidence_score:
            red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            sheet.cell(row=index, column=1).fill = red
        elif woreda_prediction[0][1] < xmed_confidence_score:
            orange = PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid')
            sheet.cell(row=index, column=1).fill = orange

        index += 1

    # Save the workbook to a BytesIO stream and send it back as an Anvil File
    output_stream = BytesIO()
    dataframe.save(output_stream)
    output_stream.seek(0)

    # Create a new Anvil File object with the modified Excel content
    output_file = anvil.BlobMedia('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', output_stream.read(), output_file_name)

    return output_file  # This will return the file to Anvil to download

# Remove the anvil.server.wait_forever() call; it's not needed in a Server Module.
